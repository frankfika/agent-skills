#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Expense reimbursement helper.

This script turns an extracted invoice manifest into a grouped reimbursement
folder plus CSV/Markdown/XLSX reports. It deliberately does not perform OCR;
the agent or user fills the manifest, and the script handles deterministic
organization, report generation, and validation.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


CATEGORIES = ["打车票", "火车飞机票", "住宿费", "餐费", "其他", "待人工确认"]
REPORT_CSV = "报销明细.csv"
REPORT_MD = "报销汇总.md"
REPORT_XLSX = "报销统计.xlsx"
MANIFEST_COLUMNS = [
    "source_path",
    "category",
    "amount",
    "tax",
    "invoice_number",
    "merchant",
    "buyer",
    "invoice_date",
    "service_date",
    "role",
    "counted",
    "note",
    "group_key",
]


@dataclass
class ExpenseRow:
    source_path: str
    category: str
    amount: Decimal
    tax: Decimal
    invoice_number: str
    merchant: str
    buyer: str
    invoice_date: str
    service_date: str
    role: str
    counted: bool
    note: str
    group_key: str
    output_path: Path | None = None

    @property
    def date_for_name(self) -> str:
        return self.service_date or self.invoice_date or "未知日期"

    @property
    def merchant_for_name(self) -> str:
        return self.merchant or "未知商户"

    @property
    def amount_text(self) -> str:
        return money(self.amount)

    @property
    def pretax(self) -> Decimal:
        return self.amount - self.tax


def money(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def parse_money(value: str) -> Decimal:
    value = (value or "0").strip().replace(",", "").replace("¥", "").replace("元", "")
    try:
        return Decimal(value or "0")
    except InvalidOperation as exc:
        raise ValueError(f"Invalid amount: {value!r}") from exc


def parse_bool(value: str) -> bool:
    return (value or "").strip().lower() in {"yes", "y", "true", "1", "是", "计入"}


def sanitize(value: str, max_len: int = 36) -> str:
    value = (value or "未知").strip()
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", "", value)
    return value[:max_len] or "未知"


def read_manifest(path: Path) -> list[ExpenseRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        missing = [col for col in MANIFEST_COLUMNS if col not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(f"Manifest missing columns: {', '.join(missing)}")

        rows: list[ExpenseRow] = []
        for index, raw in enumerate(reader, start=2):
            if not any((raw.get(col) or "").strip() for col in MANIFEST_COLUMNS):
                continue
            category = (raw.get("category") or "待人工确认").strip()
            if category not in CATEGORIES:
                category = "待人工确认"
            role = (raw.get("role") or "凭证").strip()
            if role not in {"发票", "凭证", "发票副本"}:
                role = "凭证"
            try:
                rows.append(
                    ExpenseRow(
                        source_path=(raw.get("source_path") or "").strip(),
                        category=category,
                        amount=parse_money(raw.get("amount") or "0"),
                        tax=parse_money(raw.get("tax") or "0"),
                        invoice_number=(raw.get("invoice_number") or "").strip(),
                        merchant=(raw.get("merchant") or "").strip(),
                        buyer=(raw.get("buyer") or "").strip(),
                        invoice_date=(raw.get("invoice_date") or "").strip(),
                        service_date=(raw.get("service_date") or "").strip(),
                        role=role,
                        counted=parse_bool(raw.get("counted") or ""),
                        note=(raw.get("note") or "").strip(),
                        group_key=(raw.get("group_key") or "").strip(),
                    )
                )
            except ValueError as exc:
                raise ValueError(f"Manifest row {index}: {exc}") from exc
    return rows


def init_workspace(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    (path / "raw").mkdir(exist_ok=True)
    for category in CATEGORIES:
        (path / "organized" / category).mkdir(parents=True, exist_ok=True)
    manifest = path / "manifest_template.csv"
    if not manifest.exists():
        with manifest.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(MANIFEST_COLUMNS)
            writer.writerow(
                [
                    "raw/example.pdf",
                    "打车票",
                    "36.70",
                    "0.00",
                    "1234567890",
                    "某出行平台",
                    "某公司",
                    "2026-05-25",
                    "2026-05-25",
                    "发票",
                    "yes",
                    "",
                    "",
                ]
            )
    print(f"Initialized expense workspace: {path}")


def group_id(row: ExpenseRow) -> str:
    if row.group_key:
        return row.group_key
    return f"{row.category}|{row.date_for_name}|{row.merchant_for_name}|{row.amount_text}"


def group_folder_name(rows: list[ExpenseRow]) -> str:
    primary = next((row for row in rows if row.role == "发票"), rows[0])
    return f"{sanitize(primary.date_for_name)}_{sanitize(primary.merchant_for_name)}_{primary.amount_text}元"


def source_for(row: ExpenseRow, input_root: Path) -> Path:
    src = Path(row.source_path).expanduser()
    if not src.is_absolute():
        src = input_root / src
    return src


def extension_for(path: Path) -> str:
    return path.suffix.lower() or ".pdf"


def organize(args: argparse.Namespace) -> None:
    manifest = Path(args.manifest).expanduser().resolve()
    input_root = Path(args.input_root).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    for category in CATEGORIES:
        (output_dir / category).mkdir(parents=True, exist_ok=True)

    rows = read_manifest(manifest)
    groups: dict[str, list[ExpenseRow]] = defaultdict(list)
    for row in rows:
        groups[group_id(row)].append(row)

    missing_sources: list[str] = []
    for group_rows in groups.values():
        folder = output_dir / group_rows[0].category / group_folder_name(group_rows)
        folder.mkdir(parents=True, exist_ok=True)
        for seq, row in enumerate(sorted(group_rows, key=lambda r: (r.role != "凭证", r.role, r.source_path)), start=1):
            src = source_for(row, input_root)
            if not src.exists():
                missing_sources.append(str(src))
                continue
            note = f"_{sanitize(row.note, 18)}" if row.note else ""
            filename = (
                f"{seq:02d}_{sanitize(row.date_for_name)}_{sanitize(row.role)}_"
                f"{sanitize(row.merchant_for_name)}_{row.amount_text}元{note}{extension_for(src)}"
            )
            dest = folder / filename
            if args.mode == "move":
                shutil.move(str(src), dest)
            else:
                shutil.copy2(src, dest)
            row.output_path = dest

    if missing_sources:
        print("Missing source files:")
        for src in missing_sources:
            print(f"  - {src}")

    write_reports(output_dir, rows)
    print(f"Organized reimbursement files: {output_dir}")


def counted_rows(rows: Iterable[ExpenseRow]) -> list[ExpenseRow]:
    return [row for row in rows if row.counted and row.role == "发票"]


def duplicate_notes(rows: Iterable[ExpenseRow]) -> list[str]:
    by_invoice: dict[str, list[ExpenseRow]] = defaultdict(list)
    for row in rows:
        if row.invoice_number:
            by_invoice[row.invoice_number].append(row)
    notes = []
    for invoice_number, matches in by_invoice.items():
        counted = [row for row in matches if row.counted and row.role == "发票"]
        if len(counted) > 1:
            notes.append(f"发票号码 {invoice_number} 有 {len(counted)} 条计入记录，请人工确认是否重复。")
    for row in rows:
        if row.note and (not row.counted or row.role != "发票"):
            notes.append(f"{row.source_path}: {row.note}")
    return notes


def write_reports(output_dir: Path, rows: list[ExpenseRow]) -> None:
    counted = counted_rows(rows)
    detail_path = output_dir / REPORT_CSV
    with detail_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "分类",
                "金额",
                "税额",
                "价税前金额",
                "开票日期",
                "服务日期",
                "发票号码",
                "商户",
                "购买方",
                "备注",
                "主发票文件路径",
            ]
        )
        for row in counted:
            writer.writerow(
                [
                    row.category,
                    money(row.amount),
                    money(row.tax),
                    money(row.pretax),
                    row.invoice_date,
                    row.service_date,
                    row.invoice_number,
                    row.merchant,
                    row.buyer,
                    row.note,
                    str(row.output_path.relative_to(output_dir) if row.output_path else ""),
                ]
            )

    totals: dict[str, Decimal] = {category: Decimal("0") for category in CATEGORIES}
    for row in counted:
        totals[row.category] += row.amount
    total = sum(totals.values(), Decimal("0"))

    md_path = output_dir / REPORT_MD
    with md_path.open("w", encoding="utf-8") as f:
        f.write("# 报销汇总\n\n")
        f.write("## 分类合计\n\n")
        f.write("| 分类 | 金额 |\n|---|---:|\n")
        for category in CATEGORIES:
            if totals[category]:
                f.write(f"| {category} | {money(totals[category])} |\n")
        f.write(f"| 总计 | {money(total)} |\n\n")

        f.write("## 明细\n\n")
        f.write("| 分类 | 服务日期 | 商户 | 发票号码 | 金额 | 文件 |\n|---|---|---|---|---:|---|\n")
        for row in counted:
            rel = str(row.output_path.relative_to(output_dir) if row.output_path else "")
            f.write(
                f"| {row.category} | {row.service_date or row.invoice_date} | {row.merchant} | "
                f"{row.invoice_number} | {money(row.amount)} | {rel} |\n"
            )

        notes = duplicate_notes(rows)
        if notes:
            f.write("\n## 未计入或需确认事项\n\n")
            for note in notes:
                f.write(f"- {note}\n")

    write_xlsx(output_dir, counted, totals, total)


def write_xlsx(output_dir: Path, rows: list[ExpenseRow], totals: dict[str, Decimal], total: Decimal) -> None:
    if not OPENPYXL_AVAILABLE:
        print("Skipped XLSX report: install openpyxl to enable 报销统计.xlsx")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(["分类", "金额"])
    for category in CATEGORIES:
        if totals[category]:
            ws.append([category, float(totals[category])])
    ws.append(["总计", float(total)])

    detail = wb.create_sheet("明细")
    detail.append(["分类", "金额", "税额", "价税前金额", "开票日期", "服务日期", "发票号码", "商户", "购买方", "备注", "主发票文件路径"])
    for row in rows:
        detail.append(
            [
                row.category,
                float(row.amount),
                float(row.tax),
                float(row.pretax),
                row.invoice_date,
                row.service_date,
                row.invoice_number,
                row.merchant,
                row.buyer,
                row.note,
                str(row.output_path.relative_to(output_dir) if row.output_path else ""),
            ]
        )

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E6F0FF")
        sheet.freeze_panes = "A2"

    wb.save(output_dir / REPORT_XLSX)


def validate(path: Path) -> int:
    errors: list[str] = []
    for filename in [REPORT_CSV, REPORT_MD]:
        if not (path / filename).exists():
            errors.append(f"Missing report: {filename}")
    if OPENPYXL_AVAILABLE and not (path / REPORT_XLSX).exists():
        errors.append(f"Missing report: {REPORT_XLSX}")

    csv_path = path / REPORT_CSV
    csv_total = Decimal("0")
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                csv_total += parse_money(row.get("金额", "0"))
                file_path = row.get("主发票文件路径", "")
                if file_path and not (path / file_path).exists():
                    errors.append(f"CSV references missing file: {file_path}")

    if OPENPYXL_AVAILABLE and (path / REPORT_XLSX).exists():
        wb = load_workbook(path / REPORT_XLSX, data_only=True)
        if "汇总" not in wb.sheetnames or "明细" not in wb.sheetnames:
            errors.append("XLSX must contain 汇总 and 明细 sheets")

    if errors:
        print("Validation failed:")
        for error in errors:
            print(f"  - {error}")
        return 1
    print(f"Validation ok. CSV total: {money(csv_total)}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Organize reimbursement files from a manifest.")
    sub = parser.add_subparsers(dest="command", required=True)

    init_cmd = sub.add_parser("init", help="Create a reimbursement workspace skeleton.")
    init_cmd.add_argument("path", help="Workspace path to create.")

    org = sub.add_parser("organize", help="Organize files and generate reports from a manifest.")
    org.add_argument("--manifest", required=True, help="Manifest CSV path.")
    org.add_argument("--input-root", default=".", help="Base directory for relative source_path values.")
    org.add_argument("--output-dir", required=True, help="Output directory.")
    org.add_argument("--mode", choices=["copy", "move"], default="copy", help="Copy or move source files.")

    val = sub.add_parser("validate", help="Validate generated reports and referenced files.")
    val.add_argument("path", help="Organized reimbursement directory.")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if args.command == "init":
        init_workspace(Path(args.path).expanduser().resolve())
        return 0
    if args.command == "organize":
        organize(args)
        return 0
    if args.command == "validate":
        return validate(Path(args.path).expanduser().resolve())
    parser.error("Unknown command")
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
